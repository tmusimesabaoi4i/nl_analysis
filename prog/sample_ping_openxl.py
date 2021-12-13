import time
import datetime
import openpyxl

packets_num = 1                     # パケットの数
byte_num = 1400                     # パケットのバイト数
ROUND = int(1e1)                    # 送信回数

OSNAME='linux'                      # プログラムを実行するOSの名前

if __name__ == '__main__':
    today = datetime.datetime.now()
    today_str = str(today.strftime('%Y_%m_%d_%H_%M'))

    # ファイル名は「OS＋実行日」
    file_nameXL = 'data'+ OSNAME+ today_str+ '.xlsx'
    wb = openpyxl.Workbook()                # エクセルを開く
    wb.save(file_nameXL)                    # 上書き保存

    ip_addr_array = ['192.168.1.38','192.168.1.39']        # 送り先のIPアドレス

    for ip_addr in ip_addr_array:

        if OSNAME=='linux':
            import class_ping_linux
            pping = class_ping_linux.Ping_Linux(packets_num, byte_num, ip_addr)
        elif OSNAME=='windows':
            import class_ping_windows
            pping = class_ping_windows.Ping_Windows(packets_num, byte_num, ip_addr)
        else: pass

        # シート名は「送り先IPアドレス」
        sheet_nameXL = 'SendTo_'+ ip_addr+ '.xlsx'
        ws = wb.create_sheet(title=sheet_nameXL)

        for i in range(0,ROUND):
            cellXL = ws.cell(row=i+1, column=1)
            ans = pping.run()
            ans = ans['AVE_BANDWIDTH']      # 値を取得する
            cellXL.value = ans              # セルに値を書き込む
            wb.save(file_nameXL)            # 上書き保存
            time.sleep(0.1)

        wb.close()                          # エクセルを閉じる
