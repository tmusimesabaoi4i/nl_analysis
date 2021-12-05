import socket
import sys
import time
import datetime
from openpyxl import Workbook

# Global variables
PORT = 50000           # Port No
BUFSIZE = int(2**12)   # Buffer size
PACKET_NUM = int(1e0)  # Number of packets sent
SLEEP_TIME = int(1e1)  # Number of sleep time

start_time = datetime.datetime.now()

# Create an Excel notebook
wb = Workbook()
dest_filename = 'data_book' + \
        str(start_time.strftime('%y_%m_%d')) + \
        '.xlsx'

# Initializing a notebook
ws1 = wb.active
ws1.title = 'empty'

ws2 = wb.create_sheet(title='Measurement Results')

ws2.cell(row=1, column=1, value='measured_time')
ws2.cell(row=1, column=2, value='execution_time')
ws2.cell(row=1, column=3, value='bandwidth')
ws2.cell(row=1, column=4, value='percent_receive')

for i in range(1,10):
    # Sleep to prevent errors
    print('sleep start')
    time.sleep(SLEEP_TIME)
    print('sleep end')

    # Error Number
    n_ok = 0
    n_error = 0

    # Get the measured time
    measured_time = datetime.datetime.now()

    # Measure the execution time
    time_sta = time.perf_counter()

    for j in range(PACKET_NUM):
        # Create a socket (SOCK_STREAM means a TCP socket)
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            # not a local environment, the IP address of the server.
            HOST = '192.168.1.38' # HOST = 'IP address of the server'

            try:
                # Connect to server
                sock.connect((HOST, PORT))
            except:
                print('Cant connect')
                sys.exit()

            data = bytes(BUFSIZE)

            # Send data
            sock.send(data)

            # Receive data from the server and shut down
            received = sock.recv(BUFSIZE)

            if (data == received) and \
                    (len(received) == len(data)):
                n_ok += 1
            else:
                n_error += 1

    # Measure the execution time
    time_end = time.perf_counter()
    execution_time = time_end - time_sta

    # Error rate
    percent_receive = float(n_ok)/float(n_ok+n_error)

    # Bandwidth calculation [Mbyte/sec]
    if execution_time != 0:
        bandwidth = PACKET_NUM*BUFSIZE*2/(execution_time*1e6)
    else:
        bandwidth = 0

    print('----------------------------')
    print(f'{measured_time} >>> TIME: {execution_time:f} [sec], '+ \
            f'BANDWIDTH: {bandwidth:f}[Mbyte/sec], '+ \
            f'RECEPTION: {percent_receive:f}')
    print('----------------------------')

    ws2.cell(row=i+1, column=1, value=measured_time)
    ws2.cell(row=i+1, column=2, value=execution_time)
    ws2.cell(row=i+1, column=3, value=bandwidth)
    ws2.cell(row=i+1, column=4, value=percent_receive)

wb.save(filename = dest_filename)
